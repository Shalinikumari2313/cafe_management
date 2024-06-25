from flask import Flask, render_template, request, redirect, url_for, flash
from openpyxl import load_workbook, Workbook

app = Flask(__name__)
app.secret_key = 'your_secret_key'

# Load menu from Excel
def load_menu():
    wb = load_workbook('menu.xlsx')
    sheet = wb.active
    menu = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row if present
        menu[row[0]] = row[1]
    return menu

# Load orders Excel
def load_orders():
    try:
        wb = load_workbook('orders.xlsx')
    except FileNotFoundError:
        wb = Workbook()
        wb.create_sheet(title="Orders")
        sheet = wb.active
        sheet.append(["Item", "Quantity"])  # Add headers
        wb.save('orders.xlsx')
    sheet = wb.active
    return wb, sheet

# Home page - display menu
@app.route('/')
def index():
    menu = load_menu()
    return render_template('index.html', menu=menu)

# Process order
@app.route('/order', methods=['POST'])
def order():
    print(request.form)  # Debugging statement
    item = request.form.get('item')
    quantity = int(request.form.get('quantity', 0))  # Default to 0 if quantity is missing

    if not item or quantity <= 0:
        flash('Invalid item or quantity', 'error')
        return redirect(url_for('index'))

    menu = load_menu()
    if item in menu:
        wb, sheet = load_orders()
        last_row = sheet.max_row + 1
        sheet.append([item, quantity])
        wb.save('orders.xlsx')
        flash(f'Order placed: {quantity} {item}(s)', 'success')
    else:
        flash(f'Item {item} not found in menu', 'error')
    
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True)
